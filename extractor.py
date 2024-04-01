'''
    extractor module of ChemAuto program.
    Functons: extracing initial structures and optimized structures of opt tasks 
    Developed by: Li Xilong
    Last Update: 2024-04-01
'''

import os
import re
import time
import pymol
import shutil
import openpyxl
import logging
from datetime import datetime
from logcreator import logged_input, logged_print
from pymol import cmd
from pymol import chempy
from openpyxl.styles import Font, Color, Alignment, NamedStyle
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

class StructExtractor():
    def __init__(self):
        self.elementDict = {1: 'H', 2: 'He', 3: 'Li', 4: 'Be', 5: 'B', 6: 'C', 7: 'N', 8: 'O', 9: 'F', 10: 'Ne', 
                    11: 'Na', 12: 'Mg', 13: 'Al', 14: 'Si', 15: 'P', 16: 'S', 17: 'Cl', 18: 'Ar', 19: 'K', 20: 'Ca', 
                    21: 'Sc', 22: 'Ti', 23: 'V', 24: 'Cr', 25: 'Mn', 26: 'Fe', 27: 'Co', 28: 'Ni', 29: 'Cu', 30: 'Zn', 
                    31: 'Ga', 32: 'Ge', 33: 'As', 34: 'Se', 35: 'Br', 36: 'Kr', 37: 'Rb', 38: 'Sr', 39: 'Y', 40: 'Zr', 
                    41: 'Nb', 42: 'Mo', 43: 'Tc', 44: 'Ru', 45: 'Rh', 46: 'Pd', 47: 'Ag', 48: 'Cd', 49: 'In', 50: 'Sn', 
                    51: 'Sb', 52: 'Te', 53: 'I', 54: 'Xe', 55: 'Cs', 56: 'Ba', 57: 'La', 58: 'Ce', 59: 'Pr', 60: 'Nd', 
                    61: 'Pm', 62: 'Sm', 63: 'Eu', 64: 'Gd', 65: 'Tb', 66: 'Dy', 67: 'Ho', 68: 'Er', 69: 'Tm', 70: 'Yb', 
                    71: 'Lu', 72: 'Hf', 73: 'Ta', 74: 'W', 75: 'Re', 76: 'Os', 77: 'Ir', 78: 'Pt', 79: 'Au', 80: 'Hg', 
                    81: 'Tl', 82: 'Pb', 83: 'Bi', 84: 'Po', 85: 'At', 86: 'Rn', 87: 'Fr', 88: 'Ra', 89: 'Ac', 90: 'Th', 
                    91: 'Pa', 92: 'U', 93: 'Np', 94: 'Pu', 95: 'Am', 96: 'Cm', 97: 'Bk', 98: 'Cf', 99: 'Es', 100: 'Fm', 
                    101: 'Md', 102: 'No', 103: 'Lr', 104: 'Rf', 105: 'Db', 106: 'Sg', 107: 'Bh', 108: 'Hs', 109: 'Mt', 
                    110: 'Ds', 111: 'Rg', 112: 'Cn', 113: 'Nh', 114: 'Fl', 115: 'Mc', 116: 'Lv', 117: 'Ts', 118: 'Og'}
        self.lines = []
        self.atom_num = 0
        self.coord_start_indices = []
        self.num_struct = 50
        self.excel_file_path = None
        self.base_path = None
        self.extract_type = None

    def handle_user_input(self):
        while True:
            excel_file_path = logged_input("\nPlease select the excel file ('q' to return main menu): \n")
            excel_file_path = excel_file_path.strip(' "')
            if excel_file_path.lower() == 'q':
                logged_print("\nReturn to last menu!\n")
                return 'r'
            if not os.path.exists(excel_file_path):
                logged_print("Path does not exist, please try again!")
            elif not os.path.isfile(excel_file_path):
                logged_print("This is not a valid file path, please try again!")
            elif not excel_file_path.endswith('.xlsx'):
                logged_print("This is not a valid excel file, please try again!")
            else:
                self.excel_file_path = excel_file_path
                break

        while True:
            base_path = logged_input("\nPlease enter the directory contianing log folders\n(Press Enter to use the same directory as the Excel file):\n ")
        
            if base_path == "":
                # If the user presses Enter, use the directory where the Excel file is located
                self.base_path = os.path.dirname(self.excel_file_path)
                break
            elif base_path.lower() == 'q':
                logged_print("\nReturn to last menu!\n")
                return 'r'
            elif os.path.isdir(base_path.strip(' "')):
                # If the input is a valid directory path
                self.base_path = base_path.strip(' "')
                break  # Exit the loop upon entering a valid directory path
            else:
                logged_print("\nInvalid directory path. Please enter a valid path！\n")

        while True:
            extract_type = logged_input("\nPlease choose the extract type: \n1. opted + init structures (default)  \n2. opted structures.\nEnter 'q' to return.\nYour choice:\n")
            if extract_type == "1" or extract_type == "":
                self.extract_type = 1
                break
            elif extract_type == "2":
                self.extract_type = 2
                break
            elif extract_type.lower() == 'q':
                logged_print("\nReturn to last menu!\n")
                return 'r'
            else:
                logged_print("Invalid choice, please choose again!")
        workbook = openpyxl.load_workbook(self.excel_file_path)
        sheet = workbook.active
        max_row = sheet.max_row

        while True:
            num_structures = logged_input(f"\nPlease input the number of structures to extract (1-{max_row - 1})\n")
            if num_structures.lower() == 'q':
                logged_print("\nReturn to last menu!\n")
                return 'r'
            else:
                try:
                    self.num_struct = int(num_structures)
                    if 1 <= self.num_struct <= max_row - 1:
                        break
                    else:
                        logged_print(f"Out of range（1-{max_row - 1}），please input again!")
                except ValueError:
                    logged_print("Invalid input, please try again!")

    def copy_excel_and_process_data(self, original_excel_path):
        logged_print("\nStarting copying the Excel file...")
        dir_path, file_name = os.path.split(original_excel_path)
        file_base_name, file_extension = os.path.splitext(file_name)   
        new_file_name = file_base_name + "_struc" + file_extension
        new_excel_path = os.path.join(dir_path, new_file_name)
        try:
            shutil.copy2(original_excel_path, new_excel_path)    
        except PermissionError:
            logged_print("The file is currently open. Please close the file and try again.")
            return None, None
        new_wb = load_workbook(new_excel_path)
        new_sheet = new_wb.active
        file_names = []
        for i in range(2, self.num_struct + 2):
            cell_value = new_sheet.cell(row=i, column=2).value
            if cell_value:
                file_name = os.path.join(self.base_path, os.path.splitext(cell_value)[0])
            else:
                file_name = ""
            file_names.append(file_name)
        #print(file_names[0:])
        logged_print("Excel file copied successfully.")
        return new_excel_path, file_names
        
    def generate_xyz(self, file_names):
        start_time = time.time()
        logged_print("\nGenerating xyz files, please wait....")
        xyz_paths = []
        for folder_path in file_names:
            if not folder_path:
                xyz_paths.append({"opted": "", "init": ""})
                continue
            folder_name = os.path.basename(folder_path)
            log_file_found = False
            if os.path.isdir(folder_path):
                for file in os.listdir(folder_path):
                    if file.startswith(folder_name) and file.endswith('.log'):
                        log_file_path = os.path.join(folder_path, file)
                        
                        if self.extract_type == 1:
                            xyz_path_opted = self.extract_opted_structs(log_file_path)
                            xyz_path_init = self.extract_init_structs(log_file_path)
                            xyz_paths.append({"opted": xyz_path_opted, "init": xyz_path_init})
                        elif self.extract_type == 2:
                            xyz_path_opted = self.extract_opted_structs(log_file_path)
                            xyz_paths.append({"opted": xyz_path_opted, "init": ""})
                        log_file_found = True
                        break  

            if not log_file_found:
                logged_print(f"No .log file found in folder: {folder_path}, skipping...")
                xyz_paths.append({"opted": "", "init": ""})
        
        end_time = time.time()
        execution_time = end_time - start_time
        logged_print(f"Generate all xyz files in : {execution_time} seconds")
        return xyz_paths

    def extract_init_structs(self, log_file):
        self.line=[]
        self.coord_start_indices=[]
        current_path = os.path.dirname(log_file)
        parent_dir = os.path.dirname(current_path)
        xyz_save_path = os.path.join(parent_dir, "init_structs")
        if not os.path.exists(xyz_save_path):
            os.makedirs(xyz_save_path)
        if log_file.endswith('.log'):
            with open(log_file, 'r', encoding='utf-8') as f:
                self.lines = f.readlines()
            atom_found = False
            for i, line in enumerate(self.lines):
                if 'NAtoms=' in line and not atom_found:
                    self.atom_num = int(line.split()[1])
                    atom_found = True
                if 'Standard orientation:' in line:
                    self.coord_start_indices.append(i+5)
        else:
            raise ValueError(f"Invalid file: {log_file} or not a .log file")

        if len(self.coord_start_indices) == 0:
            #No valid coordinate starting index found, returned error.
            logged_print("UnknownError: No standard orientation found")
            return ""
        else:
            coord_string = self.lines[self.coord_start_indices[0]:self.coord_start_indices[0]+self.atom_num]
        extrCoord = ''
        for item in coord_string:
            try:
                coord = [float(item.split()[3]), float(item.split()[4]), float(item.split()[5])]
                atom_type = self.elementDict[int(item.split()[1])]
                extrCoord += ' %-2s%27.8f%14.8f%14.8f\n' % (atom_type, coord[0], coord[1], coord[2])
            except ValueError:
                return "UnknownError: Invalid coordinate format"

        xyz_file_path = os.path.join(xyz_save_path, os.path.splitext(os.path.basename(log_file))[0] + "-init.xyz")
        with open(xyz_file_path, 'w', encoding='utf-8') as f:
            f.write(str(self.atom_num) + "\n")  # number of atoms
            f.write("\n")  # Comment line or empty line
            f.write(extrCoord)  # Write the coordinate data
        return xyz_file_path

    def extract_opted_structs(self, log_file):
        self.line=[]
        self.coord_start_indices=[]
        current_path = os.path.dirname(log_file)
        parent_dir = os.path.dirname(current_path)
        xyz_save_path = os.path.join(parent_dir, "opted_structs")
        if not os.path.exists(xyz_save_path):
            os.makedirs(xyz_save_path)
        if log_file.endswith('.log'):
            with open(log_file, 'r', encoding='utf-8') as f:
                self.lines = f.readlines()
            atom_found = False
            for i, line in enumerate(self.lines):
                if 'NAtoms=' in line and not atom_found:
                    self.atom_num = int(line.split()[1])
                    atom_found = True
                if 'Standard orientation:' in line:
                    self.coord_start_indices.append(i+5)
        else:
            raise ValueError(f"Invalid file: {log_file} or not a .log file")
        last_struct_index = len(self.coord_start_indices) - 1
        if len(self.coord_start_indices) == 0:
            logged_print("UnknownError: No standard orientation found")
            return ""
        if len(self.coord_start_indices) == 1:
            coord_string = self.lines[self.coord_start_indices[0]:self.coord_start_indices[0]+self.atom_num]
        else:
            # Extracting the second last structure
            coord_string = self.lines[self.coord_start_indices[last_struct_index - 1]:self.coord_start_indices[last_struct_index - 1]+self.atom_num]
            
        extrCoord = ''
        for item in coord_string:
            try:
                coord = [float(item.split()[3]), float(item.split()[4]), float(item.split()[5])]
                atom_type = self.elementDict[int(item.split()[1])]
                extrCoord += ' %-2s%27.8f%14.8f%14.8f\n' % (atom_type, coord[0], coord[1], coord[2])
            except ValueError:
                return "UnknownError: Invalid coordinate format"

        xyz_file_path = os.path.join(xyz_save_path, os.path.splitext(os.path.basename(log_file))[0] + "-opted.xyz")
        with open(xyz_file_path, 'w', encoding='utf-8') as f:
            f.write(str(self.atom_num) + "\n")  # number of atoms
            f.write("\n")  # Comment line or empty line
            f.write(extrCoord)  # Write the coordinate data
        return xyz_file_path

    def process_xyz_files(self, xyz_paths):
        start_time = time.time()
        logged_print("\nStart rendering please wait.....")
        for path_dict in xyz_paths:
            opted_path = path_dict.get("opted", "")
            init_path = path_dict.get("init", "")
            if not opted_path and not init_path:
                logged_print("Skipping opted & init xyz path: None")
            else:
                if not opted_path:
                    logged_print(f"Skipping opted xyz path: {opted_path}")
                else:
                    image_path = opted_path.replace(".xyz", ".png")
                    self.render_from_xyz(opted_path, image_path)
                
                if not init_path:
                    #print(f"Skipping init xyz path: {init_path}")
                    pass
                else:
                    image_path = init_path.replace(".xyz", ".png")
                    self.render_from_xyz(init_path, image_path)
            
        end_time = time.time()
        execution_time = end_time - start_time
        logged_print(f"Rendering time : {execution_time} seconds")

    def render_from_xyz(self, xyz_path, image_path):
        """
        Renders an image from a .xyz file using PyMOL.
        """
        cmd.reinitialize()
        # the 1st parameter of load is the file itself
        cmd.load(xyz_path, "molecule")
        cmd.hide("everything", "molecule")
        cmd.show("sticks", "molecule")
        cmd.show("spheres", "molecule")
        cmd.set("sphere_scale", 0.25, "molecule")
        cmd.set("stick_radius", 0.1, "molecule")
        cmd.set("ray_opaque_background", 1)
        cmd.set("ray_shadows", "off")
        cmd.set("specular", "off")
        cmd.set("light_count", 3)
        cmd.set("label_size", 22)
        cmd.set("label_color", "black", "molecule")
        cmd.label("molecule", "elem")  
        #cmd.zoom("molecule", complete=1)  # zoom to contain the  whole molecule
        #cmd.orient("molecule")
        #cmd.center("molecule")
        cmd.zoom("molecule", buffer=1)
        cmd.bg_color("white")
        cmd.png(image_path, width=1200, height=1200, dpi=300, ray=1)

    def insert_images_to_excel(self, new_excel_path, xyz_paths, scale_x=0.1, scale_y=0.1):
        start_time = time.time()
        
        new_wb = load_workbook(new_excel_path)
        new_sheet = new_wb.active
        # Set column width
        new_sheet.column_dimensions['D'].width = 20
        new_sheet.column_dimensions['C'].width = 20
    
        for i, path_dict in enumerate(xyz_paths, start=2):
            if self.extract_type == 1:
                # Simultaneously extract the initial structure and optimized structure
                opted_image_path = path_dict["opted"].replace(".xyz", ".png")
                init_image_path = path_dict["init"].replace(".xyz", ".png")
        
                # Handling optimized structures
                if opted_image_path:
                    if os.path.exists(opted_image_path):
                        self.insert_image_to_cell(new_sheet, 'D' + str(i), opted_image_path, scale_x, scale_y)
                    else:
                        new_sheet['D' + str(i)] = "UnknownError"
                else:
                    new_sheet['D' + str(i)] = "UnknownError"
        
                # Handling initial structures
                if init_image_path:
                    if os.path.exists(init_image_path):
                        self.insert_image_to_cell(new_sheet, 'C' + str(i), init_image_path, scale_x, scale_y)
                    else:
                        new_sheet['C' + str(i)] = "UnknownError"
                else:
                    new_sheet['C' + str(i)] = "UnknownError"
            elif self.extract_type == 2:
                opted_image_path = path_dict["opted"].replace(".xyz", ".png")
                if opted_image_path and os.path.exists(opted_image_path):
                    self.insert_image_to_cell(new_sheet, 'D' + str(i), opted_image_path, scale_x, scale_y)
                else:
                    new_sheet['D' + str(i)] = "UnknownError"
        new_wb.save(new_excel_path)
        end_time = time.time()
        execution_time = end_time - start_time
        logged_print(f"Inserting time : {execution_time} seconds\n")

    def insert_image_to_cell(self, sheet, cell, image_path, scale_x, scale_y):
        img = Image(image_path)
        img.width, img.height = img.width * scale_x, img.height * scale_y
        sheet.add_image(img, cell)
        sheet.row_dimensions[int(cell[1:])].height = 100  # Update row height

    def run(self):
        start_time = datetime.now() 
        logging.info(f"Func 2 started at {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        
        while True:
            user_input_result = self.handle_user_input()
            if user_input_result == 'r':
                return
            new_excel_path, file_names = self.copy_excel_and_process_data(self.excel_file_path)
            if new_excel_path is None or file_names is None:
                continue
            xyz_paths = self.generate_xyz(file_names)
            self.process_xyz_files(xyz_paths)
            self.insert_images_to_excel(new_excel_path, xyz_paths)
            break
        
        end_time = datetime.now()
        logging.info(f"Func 2 ended at{end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        duration_seconds = int((end_time - start_time).total_seconds())
        minutes, seconds = divmod(duration_seconds, 60)
        logging.info(f"Total duration: {minutes} minutes {seconds} seconds")
    
if __name__ == "__main__":
    extractor = StructExtractor()
    extractor.run()
