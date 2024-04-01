
'''
    analyzer module of ChemAuto program for analyzing Gaussian computational opt and sp tasks 
    Functions: extracting information such as State, HF (Hartree-Fock energy), ZPE (Zero Point Energy), Etot (Total Energy), and NImag (Number of Imaginary Frequencies).
    Features：
        4 optput formats:
                            opt+excel
                            opt+text
                            sp+excel
                            sp+text
    Developed by: Li Xilong
    Last update: 2024-04-01

'''

'''
    Maximize code reuse, although customization in Excel spreadsheets makes it challenging.
    Consolidate analysis and save types, providing four options:
        1. opt+excel 2. opt+text 3. sp+excel 4. sp+text
    Program workflow:
        1. Select analysis and save type.
        2. Choose the folder for log files (supports primary and secondary folders).
        3. Select the folder for saving results.
        4. Input the filename for saving.
        5. Execute the analysis.
    Organize code structure in the following order:
        1. Constructor method: Initializes the state of the class instance.
        2. Core process methods: Methods involving user interaction, such as starting analysis, selecting analysis and save types.
        3. Analysis execution methods: Methods that perform specific analyses, like processing different types of log files.
        4. Helper methods: Such as file creation and directory validation.
        5. Other private or specialized methods: Methods for internal use within the class.
    
    The use of == and class properties may have a slight performance advantage in theory.
    
    Regular expression matching utilizes a pattern traversal method, where the order of keywords in the pattern is crucial.
    
    Set selection branches in main().
    
    Resolve ValueError: could not convert string to float: '0.472E-01-0.207E+00' error.
        The reason was not resetting self.lines[] and self.coord_start_indices=[] for each new log processing.
    
    Set detailed timing for structure extraction.
    
    Add symmetry extraction:
        Sym regex: r"P\s*G\s*=\s*([^=\[]*)\s*\["
            Explanation: \s matches any whitespace character including space, tab, newline, etc.
                    ([^=\[]*): This is a capturing group that matches and captures any sequence of characters between = and [, excluding = and [.

    Simplify other regex: [\n\s]*---->\s*
    
    Both initial and optimized structures can be extracted, or just the optimized structures.
    
    Select the folder containing the log files separately; it can be used when the Excel file is not in the log folder.
    
    Program run test:
        For 639 output files, generating initial and optimized structures:
            Please input the num of structures to extract: 639
            Starting copying the Excel file...
            Excel file copied successfully.
            Generating xyz files, please wait...
            
            Rendering time : 238.6703372001648 seconds              ~4min
                
            Inserting time : 7.464064836502075 seconds              
            
            Total running time : 375.57376527786255 seconds         ~6.5min
            Generated Excel file size is 245Mb

        For 4011 output files, generating initial and optimized structures:
            Please input the num of structures to extract: 4011
            Starting copying the Excel file...
            Excel file copied successfully.
            Generating xyz files, please wait...
            
            Rendering time : 1501.3692154884338 seconds
            
            Inserting time : 43.93924117088318 seconds
            
            Total running time : 1861.5184030532837 seconds
    
    There is still room to improve the speed of the rendering part.
'''

import os
import re
import time
import shutil
import openpyxl
import logging
from datetime import datetime
from logcreator import setup_logging, logged_input, logged_print
from openpyxl.styles import Font, Color, Alignment, NamedStyle
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

class GaussianLogAnalyzer:
    def __init__(self):
        self.log_dir = None
        self.output_dir = None
        self.excel_save_path = None
        self.txt_save_path = None
        self.count_failed = 0
        self.count_successed = 0
        self.count_all = 0
        self.analysis_type = None
        self.save_type = None
        self.opt_patterns = {
            "sym" : r'P\s*G\s*=\s*([^=\[]*)\s*\[',
            #"state": r'S[\n\s]*t[\n\s]*a[\n\s]*t[\n\s]*e[\n\s]*=[\n\s]*([^=\\]*)[\n\s]*\\',
            "state": r'S\s*t\s*a\s*t\s*e\s*=\s*([^=\\]*)\s*\\',
            #"hf": r'H[\n\s]*F[\n\s]*=[\n\s]*([^=\\]*)[\n\s]*\\',
            "hf": r'H\s*F\s*=\s*([^=\\]*)\s*\\',
            "zpe" : r'Zero-point correction=\s*([^=\(]*)\s*\(',
            "etot" : r'Sum of electronic and zero-point Energies=\s*([^=\s]*)\n',
            #"nimag" : r'N\n*\s*I\n*\s*m\n*\s*a\n*\s*g\n*\s*=\n*\s*([^=\\]*)\\'
            "nimag" : r'N\s*I\s*m\s*a\s*g\s*=\s*([^=\\]*)\\'
        }
        self.sp_patterns = {
            #"sym" : r'P[\n\s]*G[\n\s]=[\n\s]*([^=[]*)[\n\s]*\[',
            "sym" : r'P\s*G\s*=\s*([^=[]*)\s*\[',
            #"state" : r'S[\n\s]*t[\n\s]*a[\n\s]*t[\n\s]*e[\n\s]*=[\n\s]*([^=\\]*)[\n\s]*\\',
            "state" : r'S\s*t\s*a\s*t\s*e\s*=\s*([^=\\]*)\s*\\',
            #"hf" : r'H[\n\s]*F[\n\s]*=[\n\s]*([^=\\]*)[\n\s]*\\'
            "hf" : r'H\s*F\s*=\s*([^=\\]*)\s*\\'
        }
        self.sym_patterns = {
            "C01" : "C1",
            "C02" : "C2",
            "CS" : "Cs",
            "C02H" : "C2h",
            "C02V" : "C2v",
            "C03V" : "C3v",
            "D02H" : "D2h",
            "D03H" : "D3h",
            "C04V" : "C4v",
            "D02D" : "D2d",
            "D*H" : "Dinfh",
            "C*H" : "Cinfh"
        }
    def start_analysis(self):
        choice = self.choose_analysis_type()  
        if choice == "r":  
            return  
        choice = self.setup_dirs()
        if choice == "r":  
            return  
        choice = self.create_save_file()
        if choice == "r":  
            return  
        self._perform_analysis()
        #if choice == "r":  
        #    return  
        
    def choose_analysis_type(self):
        while True:
            analysis_choice = logged_input(
                "\nPlease choose the job analysis type and save tpye:\n"
                " 1. opt + excel (default)\n"
                " 2. opt + text \n"
                " 3. sp + excel \n"
                " 4. sp + text \n"
                " 5. Return last menu\n"
                "Your choice (press Enter for default): "
            )
            
            if analysis_choice in ["1", ""]:
                logged_print("\nOPT job analysis with Excel output!")
                self.analysis_type = 'opt'
                self.save_type = 'excel'
                break
            elif analysis_choice == "2":
                logged_print("\nOPT job analysis with Text output!")
                self.analysis_type = 'opt'
                self.save_type = 'text'
                break
            elif analysis_choice == "3":
                logged_print("\nSP job analysis with Excel output!")
                self.analysis_type = 'sp'
                self.save_type = 'excel'
                break
            elif analysis_choice == "4":
                logged_print("\nSP job analysis with Text output!")
                self.analysis_type = 'sp'
                self.save_type = 'text'
                break
            elif analysis_choice in ["5", "q", "Q", "quit", "exit"]:
                logged_print("Return to last menu.\n")
                return "r"
            else:
                logged_print("\nInvalid input, please try again.")
    
    def setup_dirs(self):
        self.log_dir = self.get_valid_dir()
        if self.log_dir == "r":
            return "r"
        self.output_dir = self.create_output_dir()
        if self.output_dir == "r":  
            return "r"
    
    def get_valid_dir(self):
        while True:
            log_dir = logged_input("Please enter the directory containing .log files (or enter 'q' to quit):\n")
            log_dir = log_dir.strip(' "')
            if log_dir.lower() == 'q':
                logged_print("\nReturn to the last menu!\n")
                return "r" 
            elif os.path.isdir(log_dir):
                logged_print("\nDirectory confirmed. Proceeding with analysis.")
                return log_dir
            # Using 'break' in else would directly exit the loop; 'while True' continues indefinitely without an exit condition.
            else:
                logged_print("\nInvalid directory or directory does not exist. Please enter a valid directory path.")

    def create_output_dir(self):
        while True:
            output_dir = logged_input("Enter the output save directory (press Enter for same as log directory, or 'q' to quit): \n")
            output_dir = output_dir.strip(' "')
            if output_dir.lower() == 'q':
                logged_print("\nReturn to the last menu!\n")
                return "r" 
            elif output_dir.strip() == '':
                return self.log_dir
            elif os.path.isabs(output_dir):
                if not os.path.exists(output_dir):
                    response = logged_input("Directory doesn't exist. Input 'y' to create, or anything else to input again: ")
                    if response.strip().lower() == 'y' or response.strip() == '':
                        try:
                            os.mkdir(output_dir)
                            return output_dir
                        except OSError as e:
                            logged_print(f"\nFailed to create directory. Error: {e}\nPlease try again.")
                    else:
                        continue
                else:
                    return output_dir
            else:
                logged_print("\nInvalid input, please enter a valid path or 'q' to quit.")
    
    #actually create and overwrite the content 
    def create_save_file(self):
        save_path = self.output_dir
    
        if self.save_type == 'excel':
            while True:
                filename = logged_input("\nPlease input the name of output .xlsx file or 'q' to quit:\n")
                if filename.lower() == 'q':
                    logged_print("\nReturn to the last menu!\n")
                    return "r"
                elif filename.strip() == '':
                    logged_print("\nInvalid filename. Please try again.")
                    continue
                excel_save_path = os.path.join(save_path, filename + '.xlsx')
                logged_print(excel_save_path)
                if os.path.exists(excel_save_path):
                    logged_print("\nFile already exists.")
                    response = logged_input("\nChoose the options carefully:\n1. Overwrite it.\n2. Input a new name.\n3. Quit the program. \n\nYour choice: ")
                    if response == "1" or response.strip() == "":
                        try:
                            wb = openpyxl.Workbook()
                            wb.save(excel_save_path)
                            logged_print(f"The original file {filename}.xlsx has been overwritten successfully!\n")
                            self.excel_save_path = excel_save_path
                            break
                        except PermissionError:
                            logged_print(f"无法保存文件 '{excel_save_path}' 因为文件可能正在被其他程序使用。请关闭所有使用该文件的程序后重试。")
                    elif response == "2":
                        continue
                    elif response.lower() == "3" or response.lower() == "q":
                        logged_print("\nReturn to the last menu!\n")
                        return "r"
                    else:
                        logged_print("\nInvalid input, please try again.")
                else:
                    try:
                        wb = openpyxl.Workbook()
                        wb.save(excel_save_path)
                        logged_print(f"The output file {filename}.xlsx was created successfully!\n")
                        self.excel_save_path = excel_save_path
                        break
                    except PermissionError:
                        logged_print(f"无法保存文件 '{excel_save_path}' 因为文件可能正在被其他程序使用。请关闭所有使用该文件的程序后重试。")
        elif self.save_type == 'text':
            while True:
                filename = logged_input("Please input the name of output .txt file or 'q' to quit:\n")
                if filename.lower() == 'q':
                    logged_print("\nReturn to the last menu!\n")
                    return "r"
                elif filename.strip() == '':
                    logged_print("Invalid filename. Please try again.")
                    continue
                txt_save_path = os.path.join(save_path, filename + '.txt')
                if os.path.exists(txt_save_path):
                    logged_print("\nFile already exists!")
                    response = logged_input("Choose the options carefully:\n1. Overwrite it.\n2. Input a new name.\n3. Quit the program. \n\nYour choice: ")
                    if response == "1":
                        with open(txt_save_path, 'w', encoding='utf-8') as f:
                            pass
                        logged_print("The original file has been overwritten!!!")
                        self.txt_save_path = txt_save_path
                        break
                    elif response == "2":
                        continue
                    elif response == "3" or response == "q":
                        logged_print("\nReturn to the last menu!\n")
                        return "r"
                    else:
                        logged_print("\nInvalid input, please try again.")
                else:
                    with open(txt_save_path, 'w', encoding='utf-8') as f:
                        logged_print(f"The output file {filename}.txt was created successfully!\n")
                        self.txt_save_path = txt_save_path
                        break

    '''
    # No actual creation, only variable storage
    def create_save_file(self):
        save_path = self.output_dir
    
        if self.save_type == 'excel':
            while True:
                filename = input("\nPlease input the name of the output .xlsx file or 'q' to quit: ")
                if filename.lower() == 'q':
                    print("Program terminated!")
                    exit(0)
                elif filename.strip() == '':
                    print("Invalid filename. Please try again.")
                    continue
    
                excel_save_path = os.path.join(save_path, filename + '.xlsx')
    
                if os.path.exists(excel_save_path):
                    response = input("\nFile already exists. Choose:\n1. Overwrite\n2. Enter new name\n3. Quit\nYour choice: ")
                    if response == "1":
                        self.excel_save_path = excel_save_path
                        break
                    elif response == "2":
                        continue
                    elif response.lower() == "3":
                        print("Program terminated!")
                        exit(0)
                else:
                    self.excel_save_path = excel_save_path
                    break
    
        elif self.save_type == 'txt':
            while True:
                filename = input("\nPlease input the name of the output .txt file or 'q' to quit:")
                if filename.lower() == 'q':
                    print("Program terminated!")
                    exit(0)
                elif filename.strip() == '':
                    print("Invalid filename. Please try again.")
                    continue
    
                txt_save_path = os.path.join(save_path, filename + '.txt')
    
                if os.path.exists(txt_save_path):
                    response = input("\nFile already exists. Choose:\n1. Overwrite\n2. Enter new name\n3. Quit\nYour choice: ")
                    if response == "1":
                        self.txt_save_path = txt_save_path
                        break
                    elif response == "2":
                        continue
                    elif response == "3":
                        print("Program terminated!")
                        exit(0)
                else:
                    self.txt_save_path = txt_save_path
                    break
    '''
    
    def _perform_analysis(self):
        # Execute the corresponding analysis method according to analysis_type and save_type
        if self.analysis_type == 'opt':
            if self.save_type == 'excel':
                self.analyze_opt_logs_excel()
            elif self.save_type == 'text':
                self.analyze_opt_logs_text()
        elif self.analysis_type == 'sp':
            if self.save_type == 'excel':
                self.analyze_sp_logs_excel()
            elif self.save_type == 'text':
                self.analyze_sp_logs_text()

    def analyze_opt_logs_excel(self):
        logged_print("Analysing OPT logs, please wait.....")
        start_time = time.time()
        log_path = os.path.abspath(self.log_dir)
        log_files = [os.path.join(root, file) for root, dirs, files in os.walk(log_path) for file in files if file.endswith('.log')]
        # update total number of files
        self.count_all = len(log_files)
        # open created excel file 
        wb = openpyxl.load_workbook(self.excel_save_path)
        sheet = wb.active if wb.active else wb.create_sheet()
        sheet.append(["","Cluster","Init Struct.","Opted Struct.", "Sym.", "State", "HF", "ZPE", "Etot", "△E", "NImag"])
        for log_file in log_files:
            with open(log_file, 'r', encoding='utf-8') as file:
                log_content = file.read()
                data_row = self._process_opt_data(log_content, os.path.basename(log_file))
            sheet.append(data_row)
        self._finalize_opt_excel_sheet(sheet)
        wb.save(self.excel_save_path)
        wb.close()
        end_time = time.time()
        logged_print("OPT log analysis complete!")
        logged_print(f"Successfully analyzed {self.count_all} .log files, Normal termination: {self.count_successed}, Error termination: {self.count_failed}.")
        logged_print(f"Results are saved in {self.excel_save_path}")
        logged_print(f"Analysis completed in {end_time - start_time} seconds")

    def analyze_opt_logs_text(self):
        logged_print("Analyzing .log files, please wait.....")
        start_time = time.time()
        log_path = os.path.abspath(self.log_dir)
        log_files = [os.path.join(root, file) for root, dirs, files in os.walk(log_path) for file in files if file.endswith('.log')]
        self.count_all = len(log_files)
        with open(self.txt_save_path, 'w', encoding='utf-8') as f:
            for log_file in log_files:
                with open(log_file, 'r', encoding='utf-8') as file:
                    log_content = file.read()
                    data_row = self._process_opt_data(log_content, os.path.basename(log_file))
                    # Ensure that data_row is a string type
                    data_row_str = ' '.join([str(item) for item in data_row if item != ""])
                    f.write(data_row_str + "\n")  # Write each line of data to the file and add a newline character
        end_time = time.time()  # end timing
        logged_print(f"Successfully analyzed {self.count_all} .log files, Normal termination: {self.count_successed}, Error termination: {self.count_failed}.")
        logged_print(f"Results are saved in {self.txt_save_path}")
        logged_print(f"Analysis completed in {end_time - start_time} seconds")

    def analyze_sp_logs_excel(self):
        logged_print("Analysing SP job files, please wait.....")
        start_time = time.time()
        log_path = os.path.abspath(self.log_dir)
        log_files = [os.path.join(root, file) for root, dirs, files in os.walk(log_path) for file in files if file.endswith('.log')]
        self.count_all = len(log_files)
        wb = openpyxl.load_workbook(self.excel_save_path)
        sheet = wb.active if wb.active else wb.create_sheet()
        sheet.append(["","Cluster","Struct", "Sym.", "State", "HF", "△E"])
        
        for log_file in log_files:
            with open(log_file, 'r', encoding='utf-8') as file:
                log_content = file.read()
                data_row = self._process_sp_data(log_content, os.path.basename(log_file))
            sheet.append(data_row)
        self._finalize_sp_excel_sheet(sheet)
        wb.save(self.excel_save_path)
        wb.close()
        end_time = time.time()
        logged_print("SP job analysis complete!")
        logged_print(f"Successfully analyzed {self.count_all} .log files, Normal termination: {self.count_successed}, Error termination: {self.count_failed}.")
        logged_print(f"Results are saved in {self.excel_save_path}")
        logged_print(f"Analysis completed in {end_time - start_time} seconds")
    
    def analyze_sp_logs_text(self):
        logged_print("Analyzing .log files, please wait.....")
        start_time = time.time()
        log_path = os.path.abspath(self.log_dir)
        log_files = [os.path.join(root, file) for root, dirs, files in os.walk(log_path) for file in files if file.endswith('.log')]
        self.count_all = len(log_files)
        with open(self.txt_save_path, 'w', encoding='utf-8') as f:
            for log_file in log_files:
                with open(log_file, 'r', encoding='utf-8') as file:
                    log_content = file.read()
                    data_row = self._process_sp_data(log_content, os.path.basename(log_file))
                    data_row_str = ' '.join([str(item) for item in data_row if item != ""])
                    f.write(data_row_str + "\n")
        end_time = time.time()
        logged_print(f"Successfully analyzed {self.count_all} .log files, Normal termination: {self.count_successed}, Error termination: {self.count_failed}.")
        logged_print(f"Results are saved in {self.txt_save_path}")
        logged_print(f"Analysis completed in {end_time - start_time} seconds")

    def _process_opt_data(self, log_content, file_name):
        start_time = time.time()
        data_row = []
        data_row.append("")
        data_row.append(file_name)
        if "Normal termination of Gaussian" not in log_content:
            error_matches = re.findall(r'/g16/(.+?)\.exe', log_content)
            if error_matches:
                # Add three empty elements first, then append error information
                data_row.extend(["", "", "", error_matches[-1]])
                self.count_failed += 1
            else: 
                data_row.extend(["", "", "", "UnknownError"])
                self.count_failed += 1
        else:
            data_row.extend(["", ""])
            for key in self.opt_patterns:
                pattern = self.opt_patterns[key]
                match = re.findall(pattern, log_content)
                if key == 'sym':
                    if match:
                        content = match[-1].replace(' ', '').replace('\n', '').replace('\r', '')
                        # Find and replace the value of 'sym'
                        transformed_content = self.sym_patterns.get(content, content)   
                        data_row.append(transformed_content)
                    else:
                        data_row.append("")  
                elif key == 'state':
                    # Find and replace the value of 'state'
                    if match:
                        content = match[-1].replace(' ', '').replace('\n', '').replace('\r', '')
                        data_row.append(content)
                    else:
                        data_row.append("")
                elif key == 'hf':
                    # Find and replace the value of 'hf'
                    if match:
                        content = float(match[-1].replace(" ", "").replace("\n", ""))
                        data_row.append(content)
                    else:
                        data_row.append("")
                elif key == 'zpe':
                    # Find and replace the value of 'zpe'
                    if match:
                        content = float(match[-1])
                        data_row.append(content)
                    else:
                        data_row.append("")
                elif key == 'etot':
                    # Find and replace the value of 'etot'
                    if match:
                        content = float(match[-1])
                        data_row.append(content)
                    else:
                        data_row.append("")
                elif key == 'nimag':
                    data_row.append("")
                    # Find and replace the value of 'nimag'
                    if match:
                        content = float(match[-1].replace(' ', '').replace('\n', '').replace('\r', ''))
                        data_row.append(content)
                    else:
                        data_row.append("")
                else:
                    #  Handle other unspecified keys
                    data_row.append("")
        end_time = time.time()
        #print(f"re spend {end_time - start_time} seconds")
        return data_row

    def _process_sp_data(self, log_content, file_name):
        start_time = time.time()
        data_row = []
        data_row.append("")
        data_row.append(file_name)
        if "Normal termination of Gaussian" not in log_content:
            error_matches = re.findall(r'/g16/(.+?)\.exe', log_content)
            if error_matches:
                data_row.extend(["", "", error_matches[-1]])
                self.count_failed += 1
            else: 
                data_row.extend(["", "", "UnknownError"])
                self.count_failed += 1
        else:
            data_row.extend([""])
            for key in self.sp_patterns:
                pattern = self.opt_patterns[key]
                match = re.findall(pattern, log_content)
                if key == 'sym':
                    if match:
                        content = match[-1].replace(' ', '').replace('\n', '').replace('\r', '')
                        transformed_content = self.sym_patterns.get(content, content)
                        data_row.append(transformed_content)
                    else:
                        data_row.append("")
                elif key == 'state':
                    if match:
                        content = match[-1].replace(' ', '').replace('\n', '').replace('\r', '')
                        data_row.append(content)
                    else:
                        data_row.append("")
                elif key == 'hf':
                    if match:
                        content = float(match[-1].replace(" ", "").replace("\n", ""))
                        data_row.append(content)
                    else:
                        data_row.append("")
                else:
                    data_row.append("UnknownError")
        end_time = time.time()
        #print(f"re spend {end_time - start_time} seconds")
        #print(data_row[0:])
        return data_row

    # New private method: complete the final adjustment of the Excel spreadsheet for opt tasks
    def _finalize_opt_excel_sheet(self, sheet):
        start_time = time.time()

        self.set_sheet_format(sheet)
        self.sort_by_etot(sheet)
        self.cal_opt_diff(sheet)
        
        end_time = time.time()
        execution_time = end_time - start_time
        logged_print(f"Finalizing Excel sheet completed in {execution_time} seconds")

    # New private method: complete the final adjustment of the Excel spreadsheet for sp tasks
    def _finalize_sp_excel_sheet(self, sheet):
        start_time = time.time()
        self.set_sheet_format(sheet)
        self.sort_by_hf(sheet)
        self.cal_sp_diff(sheet)
        end_time = time.time()
        execution_time = end_time - start_time
        logged_print(f"Finalizing Excel sheet completed in {execution_time} seconds")

    def set_sheet_format(self, sheet):
        # Automatically adjust column widths
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            header_length = len(str(column_cells[0].value))#Get the length of the content of the title row
            length = max(length, header_length)  #Compare the length of the content of the title row with the length of the data rows, and take the larger value
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 6
        # Create font and alignment styles
        sheet_font = Font(name='Times New Roman', size=13)
        sheet_alignment = Alignment(horizontal='center', vertical='center')
        title_font = Font(name='Times New Roman', size=15, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')
        # Apply styles to all cells
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = sheet_font
                cell.alignment = sheet_alignment
        # Apply title styles
        for cell in sheet[1]:
            cell.font = title_font
            cell.alignment = title_alignment
        sheet.row_dimensions[1].height = 25
        
    # Sort the Etot rows in ascending order
    def sort_by_etot(self, sheet, column_index=9, descending=False):
        # Extract all the row data from the worksheet, ignoring the header row (first row)
        rows = list(sheet.iter_rows(values_only=True))
        headers, data_rows = rows[0], rows[1:]
        # Define a sort key function
        def sort_key(row):
            # get the value of the key
            value = row[column_index-1]
            # If the value is numeric, return the value; otherwise, return infinity or negative infinity depending on sort direction
            return value if isinstance(value, (int, float)) else float('inf') if not descending else float('-inf')
        # Sort all rows based on values in the specified column using the defined sort key function
        data_rows.sort(key=sort_key, reverse=descending)
        # Clear the worksheet (except for the header row)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        # Write the sorted data back into the worksheet (starting from the second row because the first row is the header row)
        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

    # Same as the previous sorting method but for the HF column
    def sort_by_hf(self, sheet, column_index=6, descending=False):
        rows = list(sheet.iter_rows(values_only=True))
        headers, data_rows = rows[0], rows[1:]
        def sort_key(row):
            value = row[column_index-1]
            return value if isinstance(value, (int, float)) else float('inf') if not descending else float('-inf')
        data_rows.sort(key=sort_key, reverse=descending)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

    def cal_opt_diff(self, sheet):
        base_value_cell = 'I2'
        multiplier = 27.21138
        column_index = 10  
        for row_idx in range(2, sheet.max_row + 1):
            # Now that column_index is defined, it can be safely used
            diff_cell = sheet.cell(row=row_idx, column = column_index)
            current_value_cell = f'I{row_idx}'
            # Calculate the difference between the current row and the base row and multiply by a constant for the opt tasks
            diff_cell.value = (
                f"=IF(AND(ISNUMBER({current_value_cell}), {current_value_cell}<>0, "
                f"ISNUMBER({base_value_cell}), {base_value_cell}<>0), "
                f"({current_value_cell}-{base_value_cell})*{multiplier}, \"\")"
                                )

    def cal_sp_diff(self, sheet):
       base_value_cell = 'F2'
       multiplier = 27.21138
       column_index = 7
       for row_idx in range(2, sheet.max_row + 1):
           diff_cell = sheet.cell(row=row_idx, column = column_index)
           current_value_cell = f'F{row_idx}'
           diff_cell.value = (
               f"=IF(AND(ISNUMBER({current_value_cell}), {current_value_cell}<>0, "
               f"ISNUMBER({base_value_cell}), {base_value_cell}<>0), "
               f"({current_value_cell}-{base_value_cell})*{multiplier}, \"\")"
                                )

    def run(self):
        start_time = datetime.now()
        logging.info(f"Func 1 started at {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        self.start_analysis()
        end_time = datetime.now()
        logging.info(f"Func 1 ended at{end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        duration_seconds = int((end_time - start_time).total_seconds())
        minutes, seconds = divmod(duration_seconds, 60)
        logging.info(f"Total duration: {minutes} minutes {seconds} seconds")

if __name__ == "__main__":
    analyzer = GaussianLogAnalyzer()
    analyzer.run()




